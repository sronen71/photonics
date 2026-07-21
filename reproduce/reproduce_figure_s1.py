#!/usr/bin/env python3
"""Reproduce and quantitatively audit Zang et al. Supplementary Fig. S1.

The physical evolution and circuit input-output relations come from this
repository's bidirectional PhCR LLE implementation.  The paper does not state
every numerical setting used for Fig. S1, so the inferred settings are kept
as named constants below and are also written to the output metrics file.

By default the script downloads the official Nature source-data workbook,
generates a paper-style reconstruction, overlays the published arrays with the
simulation, and reports objective agreement metrics.  Pass ``--paper-data``
to use an already downloaded workbook or ``--no-paper-comparison`` to run the
simulation without network access.
"""

from dataclasses import asdict, dataclass, replace
from io import BytesIO
from pathlib import Path
import argparse
import json
import sys
import urllib.request
import zipfile


REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np

from bidirectional import (
    BidirectionalParameters,
    normalized_beta_2,
    solve_bidirectional_lle,
)
from bidirectional_spectrum import bidirectional_port_power
from stochastic import GaussianModalSeed, GaussianModalWienerNoise


SOURCE_DATA_URL = (
    "https://static-content.springer.com/esm/"
    "art%3A10.1038%2Fs41566-025-01624-1/MediaObjects/"
    "41566_2025_1624_MOESM2_ESM.zip"
)
SOURCE_DATA_MEMBER = "SourceData_Supplemental_Information_S1.xlsx"

# Parameters stated in the main paper or supplement for the Fig. 1/S1 model.
CENTER_FREQUENCY_HZ = 193.1e12
QUALITY_FACTOR_INTRINSIC = 2.7e6
D2_RAD_S = -2.0 * np.pi * 8.5e6
EPSILON_PHCR = 6.0
COUPLING_FACTOR = 4.5
REFLECTOR_HALF_WIDTH = 19

# The displayed equation uses r=sqrt(R)*exp(i*phi).  With the retained
# -i*epsilon/2 coupling sign, its constructive red-mode phase is literal pi.
EQUATION_REFLECTOR_PHASE = np.pi

# These Fig. S1 choices are not all stated by the paper.  F and R were chosen
# against the official simulated source arrays, not experimental data.  The
# scan protocol controls when the subcritical soliton is seeded and accessed.
INFERRED_FORCING = 2.0
INFERRED_REFLECTIVITY = 0.90
ALPHA_MIN = 2.0
ALPHA_MAX = 10.0
SEED_ALPHA = 4.8
SPLICE_ALPHA = 4.82
CW_SCAN_RATE = 0.05
SOLITON_SCAN_RATE = 0.0125
SOLITON_SEED_RMS_AMPLITUDE = 0.02
SOLITON_SEED_MODE_WIDTH = 19.0
SOLITON_SEED_HALF_WIDTH = REFLECTOR_HALF_WIDTH
SETTLE_TIME = 30.0
SOLITON_RANDOM_SEED = 7

# The paper's exactly symmetric LLE preserves a homogeneous field forever.
# Its published alpha=4.7 spectrum nevertheless contains weak precursor
# sidebands, but the supplement does not state how they were seeded.  This
# explicitly inferred stochastic drive removes that symmetry artifact during
# the pre-soliton scan.  It is turned off before the localized branch is
# accessed, and its finite band is the paper's stated reflector band Omega.
PRECURSOR_NOISE_STRENGTH = 3.0e-5
PRECURSOR_NOISE_MODE_WIDTH = 3.0
PRECURSOR_NOISE_HALF_WIDTH = REFLECTOR_HALF_WIDTH
PRECURSOR_RANDOM_SEED = 8

# Exact display grid in the official S1 workbook.  It rounds to the paper's
# stated 200 GHz FSR and only affects the plotted optical-frequency axis.
DISPLAY_PUMP_FREQUENCY_HZ = 193.24931e12
DISPLAY_FSR_HZ = 199.38e9
SELECTED_ALPHA = (4.7, 5.35, 6.98)

PUMP_COLOR = "#e34a33"
FORWARD_COLOR = "#2ca02c"
BACKWARD_COLOR = "#3f4b9b"
DB_FLOOR = -120.0
POWER_FLOOR = 10.0 ** (DB_FLOOR / 10.0)


@dataclass(frozen=True)
class NumericalSettings:
    """Resolution controls for the standalone reconstruction."""

    spatial_points: int = 256
    time_step: float = 0.005
    low_snapshots: int = 401
    high_snapshots: int = 1201


@dataclass(frozen=True)
class FigureS1Protocol:
    """Inferred scan and perturbation settings omitted by the paper."""

    precursor_noise_strength: float = PRECURSOR_NOISE_STRENGTH
    precursor_noise_mode_width: float = PRECURSOR_NOISE_MODE_WIDTH
    precursor_noise_half_width: int = PRECURSOR_NOISE_HALF_WIDTH
    precursor_random_seed: int = PRECURSOR_RANDOM_SEED
    soliton_seed_rms_amplitude: float = SOLITON_SEED_RMS_AMPLITUDE
    soliton_seed_mode_width: float = SOLITON_SEED_MODE_WIDTH
    soliton_seed_half_width: int = SOLITON_SEED_HALF_WIDTH
    soliton_random_seed: int = SOLITON_RANDOM_SEED


def _scan(
    parameters,
    start_alpha,
    stop_alpha,
    rate,
    settings,
    snapshots,
    initial_forward,
    initial_backward,
    initial_noise=0.0,
    initial_modal_seed=None,
    modal_noise=None,
    seed=SOLITON_RANDOM_SEED,
):
    """Run one monotonic detuning segment from supplied initial fields."""
    duration = (stop_alpha - start_alpha) / rate

    def schedule(time):
        return start_alpha + rate * time

    scan_parameters = replace(parameters, alpha=stop_alpha)
    _, times, forward, backward = solve_bidirectional_lle(
        scan_parameters,
        spatial_points=settings.spatial_points,
        final_time=duration,
        time_step=settings.time_step,
        initial_noise=initial_noise,
        snapshots=snapshots,
        seed=seed,
        initial_forward=initial_forward,
        initial_backward=initial_backward,
        alpha_schedule=schedule,
        modal_noise=modal_noise,
        initial_modal_seed=initial_modal_seed,
    )
    return schedule(times), forward, backward


def _port_history(forward, backward, parameters):
    """Return shifted port spectra and their pump/comb power traces."""
    port = bidirectional_port_power(
        forward,
        backward,
        parameters,
        pump_frequency_hz=DISPLAY_PUMP_FREQUENCY_HZ,
        fsr_hz=DISPLAY_FSR_HZ,
    )
    modes = port["mode_number"]
    comb = modes != 0
    forward_power = port["forward_power_ratio"]
    backward_power = port["backward_power_ratio"]
    return {
        "frequency_thz": np.fft.fftshift(port["optical_frequency_thz"]),
        "forward_spectrum": np.fft.fftshift(forward_power, axes=1),
        "backward_spectrum": np.fft.fftshift(backward_power, axes=1),
        "forward_pump": forward_power[:, 0],
        "backward_pump": backward_power[:, 0],
        "forward_comb": np.sum(forward_power[:, comb], axis=1),
        "backward_comb": np.sum(backward_power[:, comb], axis=1),
        "intrinsic_loss": port["intrinsic_loss_ratio"],
    }


def simulate_figure_s1(settings, protocol=FigureS1Protocol()):
    """Run the CW and soliton-access segments and join them by detuning."""
    beta = normalized_beta_2(
        D2_RAD_S,
        CENTER_FREQUENCY_HZ,
        QUALITY_FACTOR_INTRINSIC,
        COUPLING_FACTOR,
    )
    parameters = BidirectionalParameters(
        alpha=ALPHA_MAX,
        forcing=INFERRED_FORCING,
        beta=beta,
        epsilon_phc=EPSILON_PHCR,
        coupling_factor=COUPLING_FACTOR,
        reflectivity=INFERRED_REFLECTIVITY,
        reflector_phase=EQUATION_REFLECTOR_PHASE,
        reflector_half_width=REFLECTOR_HALF_WIDTH,
    )

    settle_parameters = replace(parameters, alpha=ALPHA_MIN)
    _, _, settled_forward, settled_backward = solve_bidirectional_lle(
        settle_parameters,
        spatial_points=settings.spatial_points,
        final_time=SETTLE_TIME,
        time_step=settings.time_step,
        initial_noise=0.0,
        snapshots=2,
        seed=SOLITON_RANDOM_SEED,
    )
    precursor_noise = GaussianModalWienerNoise(
        strength=protocol.precursor_noise_strength,
        mode_width=protocol.precursor_noise_mode_width,
        mode_half_width=protocol.precursor_noise_half_width,
    )
    soliton_seed = GaussianModalSeed(
        rms_amplitude=protocol.soliton_seed_rms_amplitude,
        mode_width=protocol.soliton_seed_mode_width,
        mode_half_width=protocol.soliton_seed_half_width,
    )
    low_alpha, low_forward, low_backward = _scan(
        parameters,
        ALPHA_MIN,
        SEED_ALPHA,
        CW_SCAN_RATE,
        settings,
        settings.low_snapshots,
        settled_forward[-1],
        settled_backward[-1],
        modal_noise=precursor_noise,
        seed=protocol.precursor_random_seed,
    )
    high_alpha, high_forward, high_backward = _scan(
        parameters,
        SEED_ALPHA,
        ALPHA_MAX,
        SOLITON_SCAN_RATE,
        settings,
        settings.high_snapshots,
        low_forward[-1],
        low_backward[-1],
        initial_modal_seed=soliton_seed,
        seed=protocol.soliton_random_seed,
    )

    low_port = _port_history(low_forward, low_backward, parameters)
    high_port = _port_history(high_forward, high_backward, parameters)
    low_keep = low_alpha < SPLICE_ALPHA
    high_keep = high_alpha >= SPLICE_ALPHA
    alpha = np.concatenate((low_alpha[low_keep], high_alpha[high_keep]))

    result = {
        "alpha": alpha,
        "frequency_thz": low_port["frequency_thz"],
        "parameters": parameters,
        "beta": beta,
    }
    for name in (
        "forward_spectrum",
        "backward_spectrum",
        "forward_pump",
        "backward_pump",
        "forward_comb",
        "backward_comb",
        "intrinsic_loss",
    ):
        result[name] = np.concatenate(
            (low_port[name][low_keep], high_port[name][high_keep]), axis=0
        )
    result["conversion_efficiency"] = (
        result["forward_comb"] + result["backward_comb"]
    )
    result["remaining_pump"] = (
        result["forward_pump"] + result["backward_pump"]
    )
    result["energy_balance"] = (
        result["conversion_efficiency"]
        + result["remaining_pump"]
        + result["intrinsic_loss"]
    )
    return result


def _power_db(power):
    return np.maximum(10.0 * np.log10(np.maximum(power, 1.0e-30)), DB_FLOOR)


def save_reproduction_figure(simulation, output_path):
    """Render a paper-style three-panel reconstruction of Fig. S1."""
    alpha = simulation["alpha"]
    frequency = simulation["frequency_thz"]
    forward_db = _power_db(simulation["forward_spectrum"])
    backward_db = _power_db(simulation["backward_spectrum"])

    figure = plt.figure(figsize=(11.4, 8.3), constrained_layout=False)
    outer = figure.add_gridspec(
        2,
        2,
        left=0.08,
        right=0.93,
        bottom=0.09,
        top=0.95,
        width_ratios=(1.05, 1.0),
        height_ratios=(1.10, 1.0),
        wspace=0.24,
        hspace=0.30,
    )
    power_grid = outer[0, 0].subgridspec(2, 1, hspace=0.34)
    map_grid = outer[0, 1].subgridspec(2, 1, hspace=0.34)
    spectrum_grid = outer[1, :].subgridspec(2, 3, hspace=0.12, wspace=0.30)

    power_axes = [figure.add_subplot(power_grid[index, 0]) for index in range(2)]
    power_axes[0].plot(alpha, simulation["forward_pump"], color=PUMP_COLOR, lw=1.8)
    power_axes[0].plot(alpha, simulation["forward_comb"], color=FORWARD_COLOR, lw=1.8)
    power_axes[1].plot(alpha, simulation["backward_pump"], color=PUMP_COLOR, lw=1.8)
    power_axes[1].plot(alpha, simulation["backward_comb"], color=BACKWARD_COLOR, lw=1.8)
    for axis in power_axes:
        axis.set_xlim(3.0, 9.0)
        axis.set_ylim(0.0, 1.0)
        axis.set_xticks(np.arange(3, 10))
        axis.set_yticks(np.linspace(0.0, 1.0, 6))
        axis.set_xlabel(r"$\alpha$ (half-linewidths)")
        axis.set_ylabel(r"Power/$P_{\rm pump}$")
        axis.tick_params(direction="in", top=True, right=True)
    power_axes[0].text(
        0.98, 0.88, "forward", transform=power_axes[0].transAxes, ha="right"
    )
    power_axes[1].text(
        0.98, 0.88, "backward", transform=power_axes[1].transAxes, ha="right"
    )

    frequency_mask = (frequency >= 183.0) & (frequency <= 203.0)
    map_axes = [figure.add_subplot(map_grid[index, 0]) for index in range(2)]
    images = []
    for axis, values, direction in zip(
        map_axes,
        (forward_db, backward_db),
        ("forward", "backward"),
    ):
        image = axis.pcolormesh(
            alpha,
            frequency[frequency_mask],
            values[:, frequency_mask].T,
            shading="auto",
            cmap="viridis",
            vmin=DB_FLOOR,
            vmax=-10.0,
            rasterized=True,
        )
        images.append(image)
        axis.set_xlim(2.0, 10.0)
        axis.set_ylim(183.0, 203.0)
        axis.set_xlabel(r"$\alpha$ (half-linewidths)")
        axis.set_ylabel("Optical frequency (THz)")
        axis.text(
            0.98,
            0.88,
            direction,
            color="white",
            transform=axis.transAxes,
            ha="right",
        )
        axis.tick_params(direction="in", top=True, right=True)
    colorbar = figure.colorbar(
        images[-1], ax=map_axes, fraction=0.046, pad=0.025, aspect=28
    )
    colorbar.set_label("Normalized power (dB)")
    colorbar.set_ticks((-120, -100, -80, -60, -40, -20))

    spectrum_axes = np.empty((2, 3), dtype=object)
    for column, target_alpha in enumerate(SELECTED_ALPHA):
        index = int(np.argmin(abs(alpha - target_alpha)))
        for row, (name, color) in enumerate(
            (("forward_spectrum", FORWARD_COLOR), ("backward_spectrum", BACKWARD_COLOR))
        ):
            axis = figure.add_subplot(spectrum_grid[row, column])
            spectrum_axes[row, column] = axis
            values = _power_db(simulation[name][index])
            visible = (frequency >= 184.5) & (frequency <= 202.0)
            axis.vlines(
                frequency[visible],
                -100.0,
                values[visible],
                color=color,
                linewidth=0.72,
            )
            axis.set_xlim(184.5, 202.0)
            axis.set_ylim(-100.0, 5.0)
            axis.set_yticks((-100, -50, 0))
            axis.tick_params(direction="in", top=True, right=True)
            if row == 0:
                axis.tick_params(labelbottom=False)
                axis.set_title(rf"$\alpha={target_alpha:g}$", fontsize=10, pad=3)
            else:
                axis.set_xlabel("Optical frequency (THz)")
            axis.text(
                0.98,
                0.84,
                "forward" if row == 0 else "backward",
                transform=axis.transAxes,
                ha="right",
                fontsize=8,
            )

    for axis, label, x_position in (
        (power_axes[0], "a", -0.17),
        (map_axes[0], "b", -0.20),
        (spectrum_axes[0, 0], "c", -0.30),
    ):
        axis.text(
            x_position,
            1.08,
            label,
            transform=axis.transAxes,
            fontsize=18,
            fontweight="bold",
        )
    figure.text(
        0.045,
        0.275,
        "Normalized power (dB)",
        rotation=90,
        va="center",
        ha="center",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=250, facecolor="white")
    plt.close(figure)


def resolve_paper_data(requested_path, output_directory, allow_download):
    """Return the official S1 workbook path, downloading it when requested."""
    if requested_path is not None:
        path = Path(requested_path)
        if not path.is_file():
            raise FileNotFoundError(f"paper source-data workbook not found: {path}")
        return path
    if not allow_download:
        return None

    cache_directory = output_directory / "paper_source_data"
    workbook_path = cache_directory / SOURCE_DATA_MEMBER
    if workbook_path.is_file():
        return workbook_path
    cache_directory.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(
        SOURCE_DATA_URL,
        headers={"User-Agent": "photonics-figure-s1-reproduction/1.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        archive_bytes = response.read()
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        workbook_path.write_bytes(archive.read(SOURCE_DATA_MEMBER))
    return workbook_path


def load_paper_source_data(workbook_path):
    """Load the published numerical arrays for all three S1 panels."""
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "paper comparison requires openpyxl; install requirements.txt"
        ) from error

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    power_rows = list(
        workbook["Fig. 1a"].iter_rows(
            min_row=2, max_row=1001, min_col=1, max_col=6, values_only=True
        )
    )
    power = np.asarray(power_rows, dtype=float)

    maps = {}
    for sheet_name, label in (
        ("Fig. 1b-upper", "forward"),
        ("Fig. 1b-lower", "backward"),
    ):
        rows = workbook[sheet_name].iter_rows(
            min_row=2,
            max_row=1001,
            min_col=1,
            max_col=1002,
            values_only=True,
        )
        alpha = []
        frequency = []
        selected_rows = []
        for row in rows:
            alpha.append(float(row[0]))
            row_frequency = (
                float(row[1]) if row[1] is not None else float("nan")
            )
            frequency.append(row_frequency)
            if 183.0 <= row_frequency <= 203.0:
                selected_rows.append(np.asarray(row[2:], dtype=float))
        frequency = np.asarray(frequency)
        selected_frequency = frequency[
            (frequency >= 183.0) & (frequency <= 203.0)
        ]
        maps[label] = {
            "alpha": np.asarray(alpha),
            "frequency": selected_frequency,
            "power_db": np.asarray(selected_rows),
        }

    spectrum_rows = list(
        workbook["Fig. 1c"].iter_rows(
            min_row=3, max_row=514, min_col=1, max_col=12, values_only=True
        )
    )
    spectra_array = np.asarray(spectrum_rows, dtype=float)
    spectra = {}
    for target_alpha, offset in zip(SELECTED_ALPHA, (0, 4, 8)):
        spectra[str(target_alpha)] = {
            "frequency": spectra_array[:, offset],
            "forward_db": spectra_array[:, offset + 1],
            "backward_db": spectra_array[:, offset + 3],
        }
    workbook.close()
    return {"power": power, "maps": maps, "spectra": spectra}


def _threshold_interval(alpha, values, threshold):
    indices = np.flatnonzero(values >= threshold)
    if not indices.size:
        return float("nan"), float("nan")
    return float(alpha[indices[0]]), float(alpha[indices[-1]])


def compare_with_paper(simulation, paper):
    """Calculate curve, transition, map, and selected-spectrum agreement."""
    reference_power = paper["power"]
    reference_alpha = reference_power[:, 0]
    compare_mask = (reference_alpha >= 3.0) & (reference_alpha <= 9.0)
    curve_columns = {
        "forward_pump": 1,
        "forward_comb": 2,
        "backward_pump": 4,
        "backward_comb": 5,
    }
    curve_rmse = {}
    for name, column in curve_columns.items():
        interpolated = np.interp(
            reference_alpha[compare_mask],
            simulation["alpha"],
            simulation[name],
        )
        difference = interpolated - reference_power[compare_mask, column]
        curve_rmse[name] = float(np.sqrt(np.mean(difference**2)))
    overall_curve_rmse = float(
        np.sqrt(np.mean(np.square(list(curve_rmse.values()))))
    )

    paper_comb = reference_power[:, 5]
    simulation_comb = simulation["backward_comb"]
    precursor_mask = (
        (reference_alpha >= 3.8) & (reference_alpha <= 5.0)
    )
    precursor_comb = np.interp(
        reference_alpha[precursor_mask],
        simulation["alpha"],
        simulation_comb,
    )
    precursor_log_rmse = float(
        np.sqrt(
            np.mean(
                np.square(
                    np.log10(np.maximum(precursor_comb, POWER_FLOOR))
                    - np.log10(
                        np.maximum(
                            paper_comb[precursor_mask], POWER_FLOOR
                        )
                    )
                )
            )
        )
    )
    paper_peak_index = int(np.argmax(paper_comb[compare_mask]))
    paper_indices = np.flatnonzero(compare_mask)
    paper_peak_index = int(paper_indices[paper_peak_index])
    simulation_peak_mask = (
        (simulation["alpha"] >= 3.0) & (simulation["alpha"] <= 9.0)
    )
    simulation_indices = np.flatnonzero(simulation_peak_mask)
    simulation_peak_index = int(
        simulation_indices[np.argmax(simulation_comb[simulation_peak_mask])]
    )

    target = 6.98
    paper_target_index = int(np.argmin(abs(reference_alpha - target)))
    simulation_target_index = int(np.argmin(abs(simulation["alpha"] - target)))
    paper_remaining_pump = float(
        reference_power[paper_target_index, 1]
        + reference_power[paper_target_index, 4]
    )
    paper_ce_at_target = float(reference_power[paper_target_index, 5])
    simulation_ce_at_target = float(
        simulation_comb[simulation_target_index]
    )
    simulation_remaining_pump = float(
        simulation["remaining_pump"][simulation_target_index]
    )
    paper_formation, paper_collapse = _threshold_interval(
        reference_alpha[compare_mask], paper_comb[compare_mask], 0.01
    )
    simulation_formation, simulation_collapse = _threshold_interval(
        simulation["alpha"][simulation_peak_mask],
        simulation_comb[simulation_peak_mask],
        0.01,
    )

    spectrum_rmse = {}
    for target_alpha in SELECTED_ALPHA:
        reference = paper["spectra"][str(target_alpha)]
        simulation_index = int(
            np.argmin(abs(simulation["alpha"] - target_alpha))
        )
        spectrum_rmse[str(target_alpha)] = {}
        for direction in ("forward", "backward"):
            reference_db = reference[f"{direction}_db"]
            simulated_db = _power_db(
                simulation[f"{direction}_spectrum"][simulation_index]
            )
            interpolated = np.interp(
                reference["frequency"],
                simulation["frequency_thz"],
                simulated_db,
            )
            informative = (
                (reference["frequency"] >= 185.0)
                & (reference["frequency"] <= 202.0)
                & (reference_db >= -100.0)
            )
            difference = interpolated[informative] - reference_db[informative]
            spectrum_rmse[str(target_alpha)][direction] = float(
                np.sqrt(np.mean(difference**2))
            )

    map_similarity = {}
    for direction in ("forward", "backward"):
        reference = paper["maps"][direction]
        alpha_mask = (
            (reference["alpha"] >= 2.0) & (reference["alpha"] <= 10.0)
        )
        reference_map = np.clip(
            reference["power_db"][:, alpha_mask], DB_FLOOR, -10.0
        )
        simulation_db = _power_db(simulation[f"{direction}_spectrum"])
        reconstructed_rows = []
        for frequency in reference["frequency"]:
            frequency_index = int(
                np.argmin(abs(simulation["frequency_thz"] - frequency))
            )
            reconstructed_rows.append(
                np.interp(
                    reference["alpha"][alpha_mask],
                    simulation["alpha"],
                    simulation_db[:, frequency_index],
                )
            )
        reconstructed_map = np.clip(
            np.asarray(reconstructed_rows), DB_FLOOR, -10.0
        )
        informative = (reference_map > -100.0) | (reconstructed_map > -100.0)
        map_difference = reconstructed_map[informative] - reference_map[informative]
        correlation = np.corrcoef(
            reference_map[informative], reconstructed_map[informative]
        )[0, 1]
        map_similarity[direction] = {
            "rmse_db": float(np.sqrt(np.mean(map_difference**2))),
            "correlation": float(correlation),
        }

    backward_spectrum_median = float(
        np.median(
            [
                spectrum_rmse[str(alpha)]["backward"]
                for alpha in SELECTED_ALPHA
            ]
        )
    )
    metrics = {
        "curve_rmse": curve_rmse,
        "overall_curve_rmse": overall_curve_rmse,
        "precursor_backward_comb_log10_rmse": precursor_log_rmse,
        "paper_peak_backward_ce": float(paper_comb[paper_peak_index]),
        "simulation_peak_backward_ce": float(simulation_comb[simulation_peak_index]),
        "paper_peak_alpha": float(reference_alpha[paper_peak_index]),
        "simulation_peak_alpha": float(simulation["alpha"][simulation_peak_index]),
        "paper_remaining_pump_at_6_98": paper_remaining_pump,
        "simulation_remaining_pump_at_6_98": simulation_remaining_pump,
        "paper_backward_ce_at_6_98": paper_ce_at_target,
        "simulation_backward_ce_at_6_98": simulation_ce_at_target,
        "paper_formation_alpha_at_ce_0_01": paper_formation,
        "simulation_formation_alpha_at_ce_0_01": simulation_formation,
        "paper_collapse_alpha_at_ce_0_01": paper_collapse,
        "simulation_collapse_alpha_at_ce_0_01": simulation_collapse,
        "selected_spectrum_rmse_db": spectrum_rmse,
        "median_backward_spectrum_rmse_db": backward_spectrum_median,
        "spectral_map_similarity": map_similarity,
    }
    checks = {
        "peak_ce_within_0_05": abs(
            metrics["simulation_peak_backward_ce"]
            - metrics["paper_peak_backward_ce"]
        ) <= 0.05,
        "ce_at_6_98_within_0_05": abs(
            simulation_ce_at_target - paper_ce_at_target
        ) <= 0.05,
        "remaining_pump_within_0_05": abs(
            simulation_remaining_pump - paper_remaining_pump
        ) <= 0.05,
        "collapse_alpha_within_0_15": abs(
            simulation_collapse - paper_collapse
        ) <= 0.15,
        "power_curve_rmse_below_0_12": overall_curve_rmse <= 0.12,
        "backward_spectrum_rmse_below_20_db": (
            backward_spectrum_median <= 20.0
        ),
    }
    metrics["closeness_checks"] = checks
    metrics["close_to_paper"] = bool(all(checks.values()))
    return metrics


def save_comparison_figure(simulation, paper, metrics, output_path):
    """Overlay the reconstruction and official arrays for visual auditing."""
    figure = plt.figure(figsize=(12.0, 7.0), constrained_layout=True)
    grid = figure.add_gridspec(2, 3, height_ratios=(1.0, 1.05))
    forward_axis = figure.add_subplot(grid[0, 0])
    backward_axis = figure.add_subplot(grid[0, 1])
    summary_axis = figure.add_subplot(grid[0, 2])
    reference_power = paper["power"]

    for axis, prefix, color in (
        (forward_axis, "forward", FORWARD_COLOR),
        (backward_axis, "backward", BACKWARD_COLOR),
    ):
        reference_pump_column = 1 if prefix == "forward" else 4
        reference_comb_column = 2 if prefix == "forward" else 5
        axis.plot(
            reference_power[:, 0],
            reference_power[:, reference_pump_column],
            color=PUMP_COLOR,
            ls="--",
            lw=1.5,
        )
        axis.plot(
            simulation["alpha"],
            simulation[f"{prefix}_pump"],
            color=PUMP_COLOR,
            lw=1.7,
        )
        axis.plot(
            reference_power[:, 0],
            reference_power[:, reference_comb_column],
            color=color,
            ls="--",
            lw=1.5,
        )
        axis.plot(
            simulation["alpha"],
            simulation[f"{prefix}_comb"],
            color=color,
            lw=1.7,
        )
        axis.set(
            xlim=(3.0, 9.0),
            ylim=(0.0, 1.0),
            xlabel=r"$\alpha$ (half-linewidths)",
            ylabel=r"Power/$P_{\rm pump}$",
            title=f"{prefix.capitalize()} port",
        )
        axis.grid(alpha=0.2)
    forward_axis.legend(
        handles=(
            Line2D([], [], color="black", lw=1.7, label="simulation"),
            Line2D([], [], color="black", lw=1.5, ls="--", label="paper source data"),
            Line2D([], [], color=PUMP_COLOR, lw=1.7, label="pump"),
            Line2D([], [], color=FORWARD_COLOR, lw=1.7, label="comb"),
        ),
        fontsize=8,
        loc="upper left",
    )

    summary_axis.axis("off")
    verdict = "CLOSE" if metrics["close_to_paper"] else "NOT WITHIN ALL LIMITS"
    summary_lines = [
        f"Quantitative verdict: {verdict}",
        "",
        f"Backward CE at 6.98: {metrics['simulation_backward_ce_at_6_98']:.3f} "
        f"(paper {metrics['paper_backward_ce_at_6_98']:.3f})",
        f"Peak backward CE: {metrics['simulation_peak_backward_ce']:.3f} "
        f"(paper {metrics['paper_peak_backward_ce']:.3f})",
        f"Remaining pump at 6.98: "
        f"{metrics['simulation_remaining_pump_at_6_98']:.3f} "
        f"(paper {metrics['paper_remaining_pump_at_6_98']:.3f})",
        f"Collapse alpha: {metrics['simulation_collapse_alpha_at_ce_0_01']:.3f} "
        f"(paper {metrics['paper_collapse_alpha_at_ce_0_01']:.3f})",
        f"Power-curve RMSE: {metrics['overall_curve_rmse']:.3f}",
        f"Median backward-spectrum RMSE: "
        f"{metrics['median_backward_spectrum_rmse_db']:.1f} dB",
        f"Backward RMSE at alpha=4.7: "
        f"{metrics['selected_spectrum_rmse_db']['4.7']['backward']:.1f} dB",
        f"Precursor comb log10 RMSE: "
        f"{metrics['precursor_backward_comb_log10_rmse']:.2f}",
    ]
    summary_axis.text(
        0.02,
        0.98,
        "\n".join(summary_lines),
        va="top",
        ha="left",
        family="monospace",
        fontsize=9,
    )

    for column, target_alpha in enumerate(SELECTED_ALPHA):
        axis = figure.add_subplot(grid[1, column])
        reference = paper["spectra"][str(target_alpha)]
        simulation_index = int(
            np.argmin(abs(simulation["alpha"] - target_alpha))
        )
        simulated_db = _power_db(
            simulation["backward_spectrum"][simulation_index]
        )
        visible = (
            (simulation["frequency_thz"] >= 184.5)
            & (simulation["frequency_thz"] <= 202.0)
        )
        axis.plot(
            reference["frequency"],
            np.maximum(reference["backward_db"], DB_FLOOR),
            color="black",
            ls="--",
            lw=1.1,
            label="paper",
        )
        axis.plot(
            simulation["frequency_thz"][visible],
            simulated_db[visible],
            color=BACKWARD_COLOR,
            lw=1.2,
            label="simulation",
        )
        axis.set(
            xlim=(184.5, 202.0),
            ylim=(-100.0, 5.0),
            xlabel="Optical frequency (THz)",
            ylabel="Normalized power (dB)" if column == 0 else None,
            title=rf"Backward spectrum, $\alpha={target_alpha:g}$",
        )
        axis.grid(alpha=0.18)
        if column == 0:
            axis.legend(fontsize=8, loc="lower center")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220, facecolor="white")
    plt.close(figure)


def save_outputs(
    simulation,
    settings,
    protocol,
    metrics,
    output_directory,
    source_path,
):
    """Persist reproducible arrays and a human-readable metrics report."""
    output_directory.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        output_directory / "figure_s1_simulation.npz",
        alpha=simulation["alpha"],
        frequency_thz=simulation["frequency_thz"],
        forward_spectrum=simulation["forward_spectrum"],
        backward_spectrum=simulation["backward_spectrum"],
        forward_pump=simulation["forward_pump"],
        backward_pump=simulation["backward_pump"],
        forward_comb=simulation["forward_comb"],
        backward_comb=simulation["backward_comb"],
        intrinsic_loss=simulation["intrinsic_loss"],
        energy_balance=simulation["energy_balance"],
    )
    report = {
        "paper": {
            "doi": "https://doi.org/10.1038/s41566-025-01624-1",
            "source_data_url": SOURCE_DATA_URL,
            "source_data_path": str(source_path) if source_path else None,
        },
        "documented_physics": {
            "quality_factor_intrinsic": QUALITY_FACTOR_INTRINSIC,
            "d2_rad_s": D2_RAD_S,
            "epsilon_phcr": EPSILON_PHCR,
            "coupling_factor": COUPLING_FACTOR,
            "reflector_half_width": REFLECTOR_HALF_WIDTH,
            "nominal_fsr_hz": 200.0e9,
        },
        "derived_convention": {
            "equation_reflector_phase_rad": EQUATION_REFLECTOR_PHASE,
            "paper_device_phase_rad": 0.0,
        },
        "inferred_unpublished_settings": {
            "forcing": INFERRED_FORCING,
            "forcing_squared": INFERRED_FORCING**2,
            "reflectivity": INFERRED_REFLECTIVITY,
            "cw_scan_rate": CW_SCAN_RATE,
            "soliton_scan_rate": SOLITON_SCAN_RATE,
            "seed_alpha": SEED_ALPHA,
            **asdict(protocol),
        },
        "numerics": asdict(settings),
        "metrics": metrics,
    }
    (output_directory / "figure_s1_metrics.json").write_text(
        json.dumps(report, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def parse_arguments():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-directory",
        type=Path,
        default=Path("results") / "figure_s1",
        help="directory for figures, arrays, source-data cache, and metrics",
    )
    parser.add_argument(
        "--paper-data",
        type=Path,
        help=f"path to {SOURCE_DATA_MEMBER}",
    )
    parser.add_argument(
        "--no-paper-comparison",
        action="store_true",
        help="skip source-data download and quantitative paper comparison",
    )
    parser.add_argument(
        "--spatial-points",
        type=int,
        default=256,
        help="number of azimuthal samples (default: 256)",
    )
    parser.add_argument(
        "--time-step",
        type=float,
        default=0.005,
        help="normalized split-step interval (default: 0.005)",
    )
    parser.add_argument(
        "--precursor-noise-strength",
        type=float,
        default=PRECURSOR_NOISE_STRENGTH,
        help=(
            "inferred pre-soliton modal Wiener strength "
            f"(default: {PRECURSOR_NOISE_STRENGTH:g}; use 0 to disable)"
        ),
    )
    parser.add_argument(
        "--precursor-noise-mode-width",
        type=float,
        default=PRECURSOR_NOISE_MODE_WIDTH,
        help=(
            "Gaussian modal width of the inferred precursor drive "
            f"(default: {PRECURSOR_NOISE_MODE_WIDTH:g})"
        ),
    )
    parser.add_argument(
        "--soliton-seed-rms-amplitude",
        type=float,
        default=SOLITON_SEED_RMS_AMPLITUDE,
        help=(
            "grid-normalized RMS modal seed used to access the soliton branch "
            f"(default: {SOLITON_SEED_RMS_AMPLITUDE:g})"
        ),
    )
    parser.add_argument(
        "--soliton-seed-mode-width",
        type=float,
        default=SOLITON_SEED_MODE_WIDTH,
        help=(
            "Gaussian modal width of the soliton-access seed "
            f"(default: {SOLITON_SEED_MODE_WIDTH:g})"
        ),
    )
    return parser.parse_args()


def main():
    arguments = parse_arguments()
    if arguments.spatial_points < 64 or arguments.spatial_points % 2:
        raise ValueError("spatial-points must be an even integer of at least 64")
    if arguments.time_step <= 0.0:
        raise ValueError("time-step must be positive")
    if arguments.precursor_noise_strength < 0.0:
        raise ValueError("precursor-noise-strength must be nonnegative")
    if arguments.precursor_noise_mode_width <= 0.0:
        raise ValueError("precursor-noise-mode-width must be positive")
    if arguments.soliton_seed_rms_amplitude < 0.0:
        raise ValueError("soliton-seed-rms-amplitude must be nonnegative")
    if arguments.soliton_seed_mode_width <= 0.0:
        raise ValueError("soliton-seed-mode-width must be positive")
    settings = NumericalSettings(
        spatial_points=arguments.spatial_points,
        time_step=arguments.time_step,
    )
    protocol = FigureS1Protocol(
        precursor_noise_strength=arguments.precursor_noise_strength,
        precursor_noise_mode_width=arguments.precursor_noise_mode_width,
        soliton_seed_rms_amplitude=arguments.soliton_seed_rms_amplitude,
        soliton_seed_mode_width=arguments.soliton_seed_mode_width,
    )

    print("Running the bidirectional LLE scan for Figure S1...")
    simulation = simulate_figure_s1(settings, protocol)
    reproduction_path = arguments.output_directory / "figure_s1_reproduction.png"
    save_reproduction_figure(simulation, reproduction_path)

    source_path = resolve_paper_data(
        arguments.paper_data,
        arguments.output_directory,
        allow_download=not arguments.no_paper_comparison,
    )
    metrics = None
    if source_path is not None:
        print(f"Comparing against official source data: {source_path}")
        paper = load_paper_source_data(source_path)
        metrics = compare_with_paper(simulation, paper)
        save_comparison_figure(
            simulation,
            paper,
            metrics,
            arguments.output_directory / "figure_s1_comparison.png",
        )
    save_outputs(
        simulation,
        settings,
        protocol,
        metrics,
        arguments.output_directory,
        source_path,
    )

    print(f"Reproduction: {reproduction_path}")
    if metrics is None:
        print("Paper comparison skipped.")
        return
    print(
        "Backward CE at alpha=6.98: "
        f"simulation={metrics['simulation_backward_ce_at_6_98']:.6f}, "
        f"paper={metrics['paper_backward_ce_at_6_98']:.6f}"
    )
    print(
        "Remaining pump at alpha=6.98: "
        f"simulation={metrics['simulation_remaining_pump_at_6_98']:.6f}, "
        f"paper={metrics['paper_remaining_pump_at_6_98']:.6f}"
    )
    print(
        "Comb collapse alpha: "
        f"simulation={metrics['simulation_collapse_alpha_at_ce_0_01']:.4f}, "
        f"paper={metrics['paper_collapse_alpha_at_ce_0_01']:.4f}"
    )
    print(f"Power-curve RMSE: {metrics['overall_curve_rmse']:.6f}")
    print(
        "Backward-spectrum RMSE at alpha=4.7: "
        f"{metrics['selected_spectrum_rmse_db']['4.7']['backward']:.6f} dB"
    )
    print(
        "Precursor backward-comb log10 RMSE: "
        f"{metrics['precursor_backward_comb_log10_rmse']:.6f}"
    )
    print(
        "Quantitative verdict: "
        + ("CLOSE" if metrics["close_to_paper"] else "NOT WITHIN ALL LIMITS")
    )
    for name, passed in metrics["closeness_checks"].items():
        print(f"  {'PASS' if passed else 'FAIL'} {name}")


if __name__ == "__main__":
    main()
